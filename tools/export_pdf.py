# -*- coding: utf-8 -*-
"""הפקת לו"ז הצל ודפי השישי לקבוצות כ-PDF/PNG מתוך קובץ השבת.

    python3 tools/export_pdf.py 2026-09-11
    python3 tools/export_pdf.py 2026-09-11 --only shadow

הנתונים נקראים מ-shabbatot/<תאריך>.xlsx — רק תאים סטטיים («שעה» ו«אירוע» ב«לוז»,
«שמות» ב«משימות»), כך שגם עריכות שנעשו בגוגל שיטס והורדו כ-xlsx משתקפות בפלט.
הרינדור נעשה ב-Chromium שמותקן בסביבה.
"""
import argparse
import base64
import html
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
import build_workbook as bw

ROOT = Path(__file__).resolve().parent.parent
FONTS = ROOT / "assets" / "fonts"
CHROME_CANDIDATES = [
    "/opt/pw-browsers/chromium-1194/chrome-linux/chrome",
    "/opt/pw-browsers/chromium_headless_shell-1194/chrome-linux/headless_shell",
    "chromium", "chromium-browser", "google-chrome",
]
DAY_ORDER = {"שישי": 0, "שבת": 1, "מוצאי שבת": 2}
SCHEDULE_DAY = {"שישי": "שישי", "שבת": "שבת", "מוצאי שבת": "שבת"}   # מוצ"ש יושב בלו"ז תחת שבת
NO_ANCHOR = "ללא עוגן"


# ---------------------------------------------------------------------------
# קריאת קובץ השבת
# ---------------------------------------------------------------------------
def cell(ws, row, col):
    value = ws.cell(row=row, column=col).value
    if isinstance(value, str):
        value = value.strip()
    return value or None


def as_time(value):
    return value.time() if isinstance(value, datetime) else value


def split_names(text):
    return [n.strip() for n in (text or "").split(",") if n.strip()]


def read_workbook(path):
    wb = load_workbook(path)
    sched, zm = wb[bw.SH_SCHED], wb[bw.SH_ZMAN]

    shabbat_date = sched[bw.SCHED_DATE].value
    if isinstance(shabbat_date, datetime):
        shabbat_date = shabbat_date.date()
    parasha = candle = havdalah = None
    for r in range(3, zm.max_row + 1):
        value = cell(zm, r, 1)
        if isinstance(value, datetime):
            value = value.date()
        if value == shabbat_date:
            parasha, candle, havdalah = cell(zm, r, 2), as_time(cell(zm, r, 3)), as_time(cell(zm, r, 4))
            break

    events, day = [], None
    for r in range(bw.SCHED_FIRST_ROW, sched.max_row + 1):
        name = cell(sched, r, bw.L_EVENT)
        if not name:
            continue
        day = cell(sched, r, bw.L_DAY) or day
        events.append({"day": day, "hour": as_time(cell(sched, r, bw.L_HOUR)), "name": name,
                       "place": cell(sched, r, bw.L_PLACE), "note": cell(sched, r, bw.L_NOTE)})

    tasks = []
    ws = wb[bw.SH_TASKS]
    for r in range(bw.TASK_FIRST_ROW, ws.max_row + 1):
        task = cell(ws, r, bw.T_TASK)
        if not task:
            continue
        tasks.append({"row": r, "day": cell(ws, r, bw.T_DAY) or "", "hour": as_time(cell(ws, r, bw.T_HOUR)),
                      "group": cell(ws, r, bw.T_GROUP), "task": task,
                      "people": cell(ws, r, bw.T_PEOPLE), "names": split_names(cell(ws, r, bw.T_NAMES)),
                      "anchor": cell(ws, r, bw.T_ANCHOR), "note": cell(ws, r, bw.T_NOTE)})

    groups = []
    ws = wb[bw.SH_GROUPS]
    for r in range(bw.GROUP_FIRST_ROW, ws.max_row + 1):
        name = cell(ws, r, bw.G_NAME)
        if name:
            groups.append({"name": name, "parent": cell(ws, r, bw.G_PARENT), "leader": cell(ws, r, bw.G_LEADER),
                           "members": split_names(cell(ws, r, bw.G_MEMBERS)),
                           "label": name.split(" · ")[-1] if " · " in name else name})

    return {"date": shabbat_date, "parasha": parasha, "candle": candle, "havdalah": havdalah,
            "events": events, "tasks": tasks, "groups": groups}


# ---------------------------------------------------------------------------
# עזרי תצוגה
# ---------------------------------------------------------------------------
def hhmm(value):
    return "{:02d}:{:02d}".format(value.hour, value.minute) if value else ""


def esc(value):
    return html.escape(str(value)) if value else ""


def hebrew_date_range(d):
    return "{}-{}.{}".format(d.day, (d + timedelta(days=1)).day, d.month)


def when_line(data):
    when = "שבת {}".format(hebrew_date_range(data["date"])) if data["date"] else "שבת"
    if data["parasha"]:
        when += " · פרשת {}".format(data["parasha"])
    if data["candle"] and data["havdalah"]:
        when += " · כניסה {} · צאה {}".format(hhmm(data["candle"]), hhmm(data["havdalah"]))
    return when


def font_face_css():
    faces = []
    for family in ("Heebo", "Rubik"):
        for subset in ("hebrew", "latin"):
            path = FONTS / "{}-{}.woff2".format(family, subset)
            if not path.exists():
                continue
            data = base64.b64encode(path.read_bytes()).decode()
            faces.append(
                "@font-face{{font-family:'{f}';font-style:normal;font-weight:100 900;"
                "src:url(data:font/woff2;base64,{d}) format('woff2');}}".format(f=family, d=data))
    return "\n".join(faces)


BASE_CSS = """
*{box-sizing:border-box;margin:0;padding:0}
@page{size:A4;margin:0}
body{font-family:'Heebo',sans-serif;direction:rtl;color:#1F2430;background:#fff;
     -webkit-print-color-adjust:exact;print-color-adjust:exact}
.sheet{width:210mm;padding:12mm 12mm 14mm}
h1{font-family:'Rubik',sans-serif;font-size:22pt;text-align:center;color:#1F2430}
.when{text-align:center;font-size:11pt;color:#5A6572;margin-top:1.5mm}
table{width:100%;border-collapse:collapse;table-layout:fixed;margin-top:5mm}
tr{page-break-inside:avoid;break-inside:avoid}
td,th{border:1px solid #1F2430;padding:2.2mm 3mm;vertical-align:middle;font-size:10.5pt;line-height:1.5}
th{background:#F1F3F6;font-family:'Rubik',sans-serif;font-weight:600}
footer{margin-top:6mm;text-align:center;font-size:8.5pt;color:#8A94A0}
"""

SHADOW_CSS = """
.page{width:210mm;padding:10mm 10mm 12mm;position:relative}
.page.fixed{height:297mm;overflow:hidden}
.inner{transform-origin:top center;transform:scale(var(--fit,1))}
td,th{padding:1.4mm 2.5mm;font-size:9.5pt;line-height:1.35}
td.event{width:46mm;text-align:center}
td.event .day{display:block;font-family:'Rubik',sans-serif;font-weight:700;font-size:13.5pt;
              text-align:right;margin-bottom:1mm}
td.event .hour{display:block;font-family:'Rubik',sans-serif;font-size:10.5pt}
td.event .name{display:block;font-size:11.5pt;font-weight:600}
td.event .place{display:block;font-size:9pt;color:#5A6572}
td.event.loose{background:#FCE3C6}
td.lines{text-align:center}
.who{font-weight:600;line-height:1.6}
.line{padding:.4mm 0}
.line b{font-family:'Rubik',sans-serif;font-weight:600}
.line .names{color:#3A4552}
.line.all .names{font-style:italic}
"""

FRIDAY_CSS = """
h1{font-size:26pt}
.lead{text-align:center;font-size:12pt;margin-top:3mm}
.members{margin:5mm auto 0;text-align:center;font-size:12pt;line-height:1.8;max-width:170mm}
.members b{display:block;font-family:'Rubik',sans-serif;font-size:10pt;color:#5A6572}
th.hour,td.hour{width:20mm;text-align:center;font-family:'Rubik',sans-serif;font-weight:700;font-size:12pt}
td.task{font-size:12pt}
th.names,td.names{width:60mm;font-size:11pt}
.empty{margin-top:8mm;text-align:center;color:#8A94A0;font-size:12pt}
"""


# ---------------------------------------------------------------------------
# לו"ז צל
# ---------------------------------------------------------------------------
def find_event(events, task):
    """האירוע שהמשימה עוגנה אליו; משימה שאין לה עוגן בלו"ז → None."""
    if not task["anchor"]:
        return None
    hits = [e for e in events if e["name"] == task["anchor"]]
    if len(hits) > 1:
        same_day = [e for e in hits if e["day"] == SCHEDULE_DAY.get(task["day"], task["day"])]
        hits = same_day or hits
    return hits[0] if hits else None


def task_line(task, members=()):
    hour = hhmm(task["hour"])
    names = ", ".join(task["names"])
    if not names and task["people"] is None and task["group"]:
        names = "כולם"
    if members and set(task["names"]) == set(members):
        names = "כולם"                     # השמות כבר בשורת הכותרת של הקבוצה
    return ('<div class="line{all}"><b>{hour}</b>{sep}{task}'
            '{names}</div>').format(
        all=" all" if names == "כולם" else "", hour=esc(hour), sep=" — " if hour else "",
        task=esc(task["task"]),
        names=' <span class="names">— {}</span>'.format(esc(names)) if names else "")


def shadow_lines(items, groups, introduced):
    """שורות הצל לאירוע אחד. שורת השמות של קבוצה מופיעה בפעם הראשונה שהיא מופיעה;
    אחר כך מספיקים השמות שליד כל משימה."""
    by_group = {g["name"]: g for g in groups}
    seen, out = [], []
    for t in items:                         # קבוצה אחת אחרי השנייה, לפי סדר ההופעה
        if t["group"] not in seen:
            seen.append(t["group"])
    for name in seen:
        group = by_group.get(name)
        members = group["members"] if group else []
        if members and name not in introduced:
            introduced.add(name)
            out.append('<div class="who">{}</div>'.format(esc(", ".join(members))))
        out += [task_line(t, members) for t in items if t["group"] == name]
    return "".join(out)


# מתאים את לו"ז הצל לעמוד אחד: מודדים כמה הוא גולש, ומכווצים במעט לפני ההדפסה.
MEASURE_SCRIPT = """<script>
(function(){
  function report(){
    var page = document.querySelector('.page');
    var inner = document.querySelector('.inner');
    var cs = getComputedStyle(page);
    var avail = page.clientHeight - parseFloat(cs.paddingTop) - parseFloat(cs.paddingBottom);
    document.title = 'FIT:' + (avail / inner.getBoundingClientRect().height);
  }
  window.addEventListener('load', report);
  if (document.fonts && document.fonts.ready) { document.fonts.ready.then(report); }
})();
</script>"""
MIN_FIT = 0.72          # מתחת לזה עדיף שני עמודים קריאים מעמוד אחד זעיר


def shadow_html(data, fit=1.0, measure=False):
    events, groups = data["events"], data["groups"]
    by_event = {id(e): [] for e in events}
    loose = []
    for t in sorted(data["tasks"], key=lambda t: (DAY_ORDER.get(t["day"], 9), t["hour"] is None,
                                                 t["hour"] or datetime.min.time(), t["row"])):
        e = find_event(events, t)
        (by_event[id(e)] if e else loose).append(t)

    rows, last_day, introduced = [], None, set()
    for e in events:
        day_head = ""
        if e["day"] != last_day:
            day_head = '<span class="day">יום {}:</span>'.format(esc(e["day"]))
            last_day = e["day"]
        rows.append('<tr><td class="event">{day}<span class="hour">{hour}</span>'
                    '<span class="name">{name}</span>{place}</td><td class="lines">{lines}</td></tr>'.format(
                        day=day_head, hour=esc(hhmm(e["hour"])), name=esc(e["name"]),
                        place='<span class="place">{}</span>'.format(esc(e["place"])) if e["place"] else "",
                        lines=shadow_lines(by_event[id(e)], groups, introduced) or "&nbsp;"))
    if loose:
        rows.append('<tr><td class="event loose"><span class="name">{}</span></td>'
                    '<td class="lines">{}</td></tr>'.format(NO_ANCHOR, shadow_lines(loose, groups, introduced)))

    fixed = measure or fit < 1.0
    return """<!doctype html><html lang="he" dir="rtl"><head><meta charset="utf-8">
<style>{fonts}{base}{css}:root{{--fit:{fit}}}</style></head><body>
<div class="page{fixed}"><div class="inner">
<h1>לו"ז שבת ולו"ז צל</h1><div class="when">{when}</div>
<table><tbody>{rows}</tbody></table>
<footer>מדרשת עין פרת</footer>
</div></div>{script}</body></html>""".format(
        fonts=font_face_css(), base=BASE_CSS, css=SHADOW_CSS, fit=fit, fixed=" fixed" if fixed else "",
        when=esc(when_line(data)), rows="".join(rows), script=MEASURE_SCRIPT if measure else "")


# ---------------------------------------------------------------------------
# דף שישי לקבוצה
# ---------------------------------------------------------------------------
def friday_html(group, tasks, data):
    title = "{} · {}".format(group["parent"], group["label"]) if group["parent"] else group["name"]
    rows = []
    for t in sorted(tasks, key=lambda t: (t["hour"] is None, t["hour"] or datetime.min.time(), t["row"])):
        names = ", ".join(t["names"]) or ("כולם" if t["people"] is None else "")
        rows.append('<tr><td class="hour">{}</td><td class="task">{}</td><td class="names">{}</td></tr>'.format(
            esc(hhmm(t["hour"])) or "—", esc(t["task"]), esc(names)))
    table = ('<table><thead><tr><th class="hour">שעה</th><th>משימה</th><th class="names">מי</th></tr></thead>'
             '<tbody>{}</tbody></table>'.format("".join(rows))) if rows else \
        '<div class="empty">אין משימות ליום שישי</div>'
    return """<!doctype html><html lang="he" dir="rtl"><head><meta charset="utf-8">
<style>{fonts}{base}{css}</style></head><body><div class="sheet">
<h1>{title}</h1><div class="when">יום שישי · {when}</div>
{lead}
<div class="members"><b>חברי הקבוצה · {n}</b>{members}</div>
{table}
<footer>מדרשת עין פרת · הכנות שישי</footer>
</div></body></html>""".format(
        fonts=font_face_css(), base=BASE_CSS, css=FRIDAY_CSS, title=esc(title), when=esc(when_line(data)),
        lead='<div class="lead">מוביל/ה: <b>{}</b></div>'.format(esc(group["leader"])) if group["leader"] else "",
        n=len(group["members"]), members=esc(", ".join(group["members"])) or "—", table=table)


# ---------------------------------------------------------------------------
# רינדור ב-Chromium
# ---------------------------------------------------------------------------
def chrome_binary():
    for candidate in CHROME_CANDIDATES:
        path = shutil.which(candidate) or (candidate if Path(candidate).exists() else None)
        if path:
            return path
    raise SystemExit("לא נמצא דפדפן Chromium להפקת ה-PDF")


def measure_fit(html_text):
    """מריץ את העמוד ב-Chromium ומחזיר את מקדם ההתאמה לעמוד אחד (1.0 = נכנס)."""
    chrome = chrome_binary()
    with tempfile.TemporaryDirectory() as tmp:
        source = Path(tmp) / "measure.html"
        source.write_text(html_text, encoding="utf-8")
        result = subprocess.run(
            [chrome, "--headless", "--disable-gpu", "--no-sandbox", "--virtual-time-budget=8000",
             "--user-data-dir={}/profile".format(tmp), "--dump-dom", source.as_uri()],
            capture_output=True, timeout=180)
        match = re.search(r"FIT:([0-9.]+)", result.stdout.decode("utf-8", "replace"))
        if not match:
            return 1.0
        ratio = float(match.group(1))
        return 1.0 if ratio >= 1.03 else ratio * 0.97


def render(html_text, out_pdf, out_png=None):
    chrome = chrome_binary()
    with tempfile.TemporaryDirectory() as tmp:
        source = Path(tmp) / "page.html"
        source.write_text(html_text, encoding="utf-8")
        common = [chrome, "--headless", "--disable-gpu", "--no-sandbox", "--hide-scrollbars",
                  "--virtual-time-budget=6000", "--user-data-dir={}/profile".format(tmp)]
        subprocess.run(common + ["--no-pdf-header-footer", "--print-to-pdf-no-header",
                                 "--print-to-pdf={}".format(out_pdf), source.as_uri()],
                       check=True, capture_output=True, timeout=180)
        if out_png:
            pages = pdf_pages(out_pdf)        # התמונה מכסה את כל העמודים, לא רק את הראשון
            subprocess.run(common + ["--window-size=794,{}".format(1123 * pages),
                                     "--force-device-scale-factor=2",
                                     "--screenshot={}".format(out_png), source.as_uri()],
                           check=True, capture_output=True, timeout=180)
    return pdf_pages(out_pdf)


def pdf_pages(path):
    return max(1, len(re.findall(rb"/Type\s*/Page[^s]", Path(path).read_bytes())))


def main():
    ap = argparse.ArgumentParser(description="הפקת לו\"ז צל ודפי שישי")
    ap.add_argument("date", help="תאריך יום שישי, למשל 2026-09-11")
    ap.add_argument("--only", choices=["friday", "shadow"], help="להפיק רק חלק")
    ap.add_argument("--no-png", action="store_true", help="בלי תמונות PNG")
    args = ap.parse_args()

    d = datetime.strptime(args.date, "%Y-%m-%d").date()
    source = ROOT / "shabbatot" / "{}.xlsx".format(d.isoformat())
    if not source.exists():
        raise SystemExit("לא נמצא {} — הריצו קודם tools/new_shabbat.py".format(source))
    data = read_workbook(source)
    out_dir = ROOT / "shabbatot" / d.isoformat()
    out_dir.mkdir(parents=True, exist_ok=True)
    for old in list(out_dir.glob("*.pdf")) + list(out_dir.glob("*.png")):
        old.unlink()                      # בלי פלטים ישנים מהרצה קודמת

    made = []
    if args.only != "shadow":
        for group in data["groups"]:
            friday = [t for t in data["tasks"] if t["group"] == group["name"] and t["day"] == "שישי"]
            if not friday:
                continue
            title = "{} · {}".format(group["parent"], group["label"]) if group["parent"] else group["name"]
            pdf = out_dir / "{}.pdf".format(title)
            png = None if args.no_png else out_dir / "{}.png".format(title)
            render(friday_html(group, friday, data), pdf, png)
            made += [x for x in (pdf, png) if x]
    if args.only != "friday":
        pdf = out_dir / "לוז צל.pdf"
        png = None if args.no_png else out_dir / "לוז צל.png"
        fit = measure_fit(shadow_html(data, measure=True))
        if fit < MIN_FIT:
            fit = 1.0                     # צפוף מדי לעמוד אחד — עדיף שני עמודים קריאים
        pages = render(shadow_html(data, fit=fit), pdf, png)
        made += [x for x in (pdf, png) if x]
        print("  לו\"ז צל: {} עמוד/ים".format(pages))

    for path in made:
        print("  {:<40} {:>8,} bytes".format(path.name, path.stat().st_size))
    print("{} קבצים ← {}/".format(len(made), out_dir.relative_to(ROOT)))


if __name__ == "__main__":
    main()
