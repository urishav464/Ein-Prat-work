# -*- coding: utf-8 -*-
"""קריאת רשימת החניכים מתיקיית data/ ונרמולה ל-data/students.csv.

מזהה אוטומטית כל קובץ .xlsx/.csv ששמו מכיל "חניכ" או "students" (למשל ייצוא של
טופס גוגל), לוקח את עמודת השמות ומייצר רשימה נקייה. אם אין קובץ — נוצרים
placeholders, כדי שהמערכת תמשיך לעבוד עד שהרשימה האמיתית תגיע.

    python3 tools/roster.py
"""
import csv
import difflib
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT = DATA / "students.csv"
HEADERS = ["שם", "תוכנית", "קבוצה קבועה", "הערה"]
PLACEHOLDER_COUNT = 40

CANONICAL = {"מישל דויד": "מישל דוד"}

NAME_HINTS = ("שם", "name", "חניך")
SKIP_HINTS = ("חותמת", "timestamp", "מייל", "email", "טלפון", "phone")


ELUL_PATTERN = re.compile(r"(elul|אלול)", re.UNICODE)


def _source_files(elul=False):
    pattern = re.compile(r"(חניכ|students|רשימ)", re.UNICODE)
    files = [p for p in DATA.glob("*")
             if p.suffix.lower() in (".xlsx", ".xlsm", ".csv")
             and p.name != OUT.name and pattern.search(p.stem)]
    return sorted(p for p in files if bool(ELUL_PATTERN.search(p.stem)) == elul)


def _pick_name_column(header, rows):
    """עמודת השמות: לפי כותרת אם אפשר, אחרת העמודה עם הכי הרבה ערכי טקסט."""
    for i, title in enumerate(header):
        t = str(title or "").strip().lower()
        if any(h in t for h in NAME_HINTS) and not any(s in t for s in SKIP_HINTS):
            return i
    best, best_score = 0, -1
    for i in range(len(header)):
        score = sum(1 for r in rows if i < len(r) and isinstance(r[i], str) and r[i].strip())
        if score > best_score:
            best, best_score = i, score
    return best


def _read_rows(path):
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.reader(fh))
        return (rows[0] if rows else []), rows[1:]
    from openpyxl import load_workbook
    ws = load_workbook(path, data_only=True).worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    return (rows[0] if rows else []), rows[1:]


def load_names(elul=False):
    for path in _source_files(elul=elul):
        try:
            header, rows = _read_rows(path)
        except Exception as exc:                       # קובץ פגום/נעול — ננסה את הבא
            print("דילוג על {}: {}".format(path.name, exc))
            continue
        if not rows:
            continue
        col = _pick_name_column(header, rows)
        names, seen = [], set()
        for row in rows:
            value = row[col] if col < len(row) else None
            name = " ".join(str(value).split()) if value else ""
            name = CANONICAL.get(name, name)
            if name and name not in seen:
                seen.add(name)
                names.append(name)
        if names:
            return names, path.name
    if elul:
        return [], None
    return ["חניך {}".format(i) for i in range(1, PLACEHOLDER_COUNT + 1)], None


def read_students():
    """הרשימה המנורמלת כ-[{שם, קבוצה קבועה, הערה}], אחרי שהורצה main()."""
    if not OUT.exists():
        main()
    with OUT.open(encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def match_name(name, roster):
    """התאמת שם שהוקלד לשם הקנוני ברשימה. מחזיר (שם, איך) או (None, "לא נמצא")."""
    table = {_norm(r): r for r in roster}
    n = _norm(CANONICAL.get(" ".join(str(name).split()), name))
    if n in table:
        return table[n], "מדויק"
    words = set(n.split())
    contained = [orig for key, orig in table.items()
                 if words <= set(key.split()) or set(key.split()) <= words]
    if len(contained) == 1:
        return contained[0], "הכלה"
    close = difflib.get_close_matches(n, list(table), n=1, cutoff=0.8)
    if close:
        return table[close[0]], "דמיון"
    return None, "לא נמצא"


def _norm(text):
    text = re.sub(r"[\u0591-\u05C7]", "", str(text))
    text = text.replace("״", '"').replace("׳", "'").replace("-", " ")
    return re.sub(r"\s+", " ", text).strip()


def main():
    names, source = load_names()
    elul_names, elul_source = load_names(elul=True)
    existing = {}
    if OUT.exists():                                   # שימור נעילות והערות קיימות
        with OUT.open(encoding="utf-8-sig", newline="") as fh:
            existing = {r["שם"]: r for r in csv.DictReader(fh)}

    with OUT.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=HEADERS)
        writer.writeheader()
        for name, program in ([(n, "מדרשה") for n in names]
                              + [(n, "אלול") for n in elul_names if n not in names]):
            prev = existing.get(name, {})
            writer.writerow({"שם": name,
                             "תוכנית": program,
                             "קבוצה קבועה": prev.get("קבוצה קבועה", ""),
                             "הערה": prev.get("הערה", "")})

    print("{} חניכי מדרשה (מקור: {}) + {} חניכי אלול (מקור: {}) ← {}".format(
        len(names), source or "placeholder", len(elul_names),
        elul_source or "—", OUT.relative_to(ROOT)))
    return names + elul_names


if __name__ == "__main__":
    sys.exit(0 if main() else 1)
