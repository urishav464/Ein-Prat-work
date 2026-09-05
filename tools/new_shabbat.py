# -*- coding: utf-8 -*-
"""הפקת קובץ עבודה לשבת מסוימת.

    python3 tools/new_shabbat.py 2026-09-11
    python3 tools/new_shabbat.py 2026-09-11 --from shabbatot/2026-09-04.xlsx

ברירת המחדל: עותק של shabbat-planner.xlsx. עם --from מתחילים מקובץ של שבת קודמת —
כך עריכות שנעשו במשימות, במתכונים ובקייטרינג עוברות הלאה; רק השמות, הקבוצות
והשעות מתאפסים. בשני המקרים התאריך נכתב ב«לוז», ועמודת «שעה» מתמלאת בערכים
סטטיים לפי data/schedule_template.csv (ההצעה בגיליון היא נוסחה, והייצוא לא
יכול לקרוא נוסחאות — לכן השעות נכתבות כאן כמספרים).
"""
import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

sys.path.insert(0, str(Path(__file__).resolve().parent))
import build_workbook as bw

ROOT = Path(__file__).resolve().parent.parent
TEMPLATE = ROOT / "shabbat-planner.xlsx"
OUTDIR = ROOT / "shabbatot"


def clear(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = None


def find_row(ws, date):
    """שורת השבת בגיליון «זמנים», או None אם התאריך אינו בלוח."""
    for row in ws.iter_rows(min_row=3, max_col=1):
        cell = row[0]
        if cell.value and getattr(cell.value, "date", lambda: cell.value)() == date:
            return cell.row
    return None


def reset_schedule(wb, date):
    zm, ws = wb[bw.SH_ZMAN], wb[bw.SH_SCHED]
    row = find_row(zm, date)
    if row is None:
        raise SystemExit("התאריך {} אינו מופיע בגיליון «זמנים». הריצו קודם tools/zmanim.py"
                         .format(date.strftime("%d/%m/%Y")))
    candle, havdalah = zm.cell(row=row, column=3).value, zm.cell(row=row, column=4).value
    ws[bw.SCHED_DATE] = zm.cell(row=row, column=1).value

    template = bw.read_schedule_template()
    by_event = {(t["יום"], t["אירוע"]): t for t in template}
    for r in range(bw.SCHED_FIRST_ROW, bw.SCHED_FIRST_ROW + bw.SCHED_ROWS):
        day, event = ws.cell(row=r, column=bw.L_DAY).value, ws.cell(row=r, column=bw.L_EVENT).value
        spec = by_event.get((day, event))
        if spec:
            ws.cell(row=r, column=bw.L_HOUR).value = bw.suggested_time(spec, candle, havdalah)
        elif event:
            clear(ws, r, bw.L_HOUR)      # אירוע שאורי הוסיף — השעה שלו נקבעת ידנית
    return zm.cell(row=row, column=2).value, candle, havdalah


def reset_people(wb):
    ws = wb[bw.SH_TASKS]
    for r in range(bw.TASK_FIRST_ROW, ws.max_row + 1):
        clear(ws, r, bw.T_NAMES)
    ws = wb[bw.SH_GROUPS]
    for r in range(bw.GROUP_FIRST_ROW, ws.max_row + 1):
        clear(ws, r, bw.G_MEMBERS)
    ws = wb[bw.SH_STUDENTS]
    for r in range(bw.STUDENT_FIRST_ROW, ws.max_row + 1):
        for col in (bw.S_AVAILABLE, bw.S_GROUP, bw.S_NOTE):
            clear(ws, r, col)


def main():
    ap = argparse.ArgumentParser(description="הפקת קובץ שבת")
    ap.add_argument("date", help="תאריך יום שישי, למשל 2026-09-11")
    ap.add_argument("--from", dest="source", help="להתחיל מקובץ של שבת קודמת במקום מהתבנית")
    ap.add_argument("--out", help="נתיב פלט חלופי")
    args = ap.parse_args()

    date = datetime.strptime(args.date, "%Y-%m-%d").date()
    if date.weekday() != 4:
        print("שים לב: {} אינו יום שישי — ממשיך בכל זאת.".format(args.date))

    OUTDIR.mkdir(exist_ok=True)
    out = Path(args.out) if args.out else OUTDIR / "{}.xlsx".format(date.isoformat())
    source = Path(args.source) if args.source else TEMPLATE
    if source.resolve() != out.resolve():
        shutil.copy(source, out)

    wb = load_workbook(out)
    parasha, candle, havdalah = reset_schedule(wb, date)
    reset_people(wb)
    wb.save(out)
    shown = out.relative_to(ROOT) if out.resolve().is_relative_to(ROOT) else out
    print("נוצר: {}  ({}, {})".format(shown, date.strftime("%d/%m/%Y"), parasha or "—"))
    print("  כניסת שבת {}  ·  צאת שבת {}".format(
        candle.strftime("%H:%M"), havdalah.strftime("%H:%M")))


if __name__ == "__main__":
    main()
