# -*- coding: utf-8 -*-
"""מחולל הגיליון המתכלל (shabbat-planner.xlsx) — מדרשת עין פרת.

    python3 tools/build_workbook.py

הקובץ נבנה מאפס מתוך data/*.csv ומיועד לייבוא לגוגל שיטס (קובץ ← ייבוא), ולכן
משתמש רק בנוסחאות בסיסיות שעובדות גם שם. השימוש השוטף לא דורש את הסקריפט —
מריצים אותו מחדש רק כשמשנים את מבנה הגיליון.

שמונה גיליונות: הוראות · לוז · משימות · קבוצות · חניכים · מתכונים · קייטרינג · זמנים.
"""
import csv
from datetime import datetime, time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT = ROOT / "shabbat-planner.xlsx"

# --- שמות גיליונות (בלי גרשיים — בטוח לנוסחאות ולשיטס) ------------------------
SH_HELP, SH_SCHED, SH_TASKS, SH_GROUPS = "הוראות", "לוז", "משימות", "קבוצות"
SH_STUDENTS, SH_RECIPES, SH_CATERING, SH_ZMAN = "חניכים", "מתכונים", "קייטרינג", "זמנים"

# --- כתובות קבועות שסקריפטים אחרים מסתמכים עליהן ------------------------------
SCHED_DATE = "B3"          # תאריך השבת (קלט)
SCHED_CANDLE = "B4"        # כניסת שבת (נוסחה מ«זמנים»)
SCHED_HAVDALAH = "D4"      # צאת שבת (נוסחה מ«זמנים»)
SCHED_HEADER_ROW = 6       # כותרת טבלת הלו"ז; השורות מתחילות מ-7
SCHED_FIRST_ROW = SCHED_HEADER_ROW + 1
SCHED_ROWS = 30            # שורות לו"ז (כולל רזרבה לאירועים שאורי מוסיף)

TASK_FIRST_ROW = 3
TASK_ROWS = 120
GROUP_FIRST_ROW = 3
GROUP_ROWS = 30
STUDENT_FIRST_ROW = 3
STUDENT_ROWS = 130
RECIPE_ROWS = 80
CATERING_ROWS = 50

# עמודות בגיליון «משימות»
T_DAY, T_HOUR, T_GROUP, T_TASK, T_PEOPLE, T_NAMES, T_ANCHOR, T_NOTE = range(1, 9)
# עמודות בגיליון «קבוצות»
G_NAME, G_PARENT, G_LEADER, G_MEMBERS, G_SIZE, G_COUNT = range(1, 7)
# עמודות בגיליון «חניכים»
S_NAME, S_PROGRAM, S_AVAILABLE, S_GROUP, S_NOTE = range(1, 6)
# עמודות בגיליון «לוז»
L_DAY, L_HOUR, L_EVENT, L_PLACE, L_NOTE, L_SUGGEST = range(1, 7)

DAYS = "שישי,שבת,מוצאי שבת"
RECIPE_KINDS = "עוגות,סלטים,מטבוחה,ארוחת צהריים שישי"
MEALS = "ארוחת ערב,קידוש,ארוחת צהריים,סעודה שלישית"

# --- צבעים ------------------------------------------------------------------
INK, MUTED, LINE, BAND = "1F2430", "6B7280", "C9CFD8", "EDF1F6"
ACCENT, INPUT_BG, CALC_BG = "2E5C8A", "FFF9E3", "EEF3F8"
TAB_INPUT, TAB_OUTPUT, TAB_REF = "E8A33D", "2E5C8A", "9AA5B1"
DAY_FILLS = {"שישי": "E8F0F8", "שבת": "F3EDE3", "מוצאי שבת": "EDEAF5"}
FONT = "Arial"


# ---------------------------------------------------------------------------
# עוזרי עיצוב
# ---------------------------------------------------------------------------
def f(size=11, bold=False, color=INK, italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)


def fill(rgb):
    return PatternFill("solid", fgColor=rgb)


def align(h="right", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, readingOrder=2)


def box(color=LINE, style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def q(sheet):
    return "'{}'!".format(sheet)


def page(ws, tab=None):
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if tab:
        ws.sheet_properties.tabColor = tab


def widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


def title_row(ws, row, text, span="A:F", size=18):
    last = span.split(":")[1]
    ws.merge_cells("A{r}:{c}{r}".format(r=row, c=last))
    c = ws["A{}".format(row)]
    c.value = text
    c.font = f(size, bold=True, color=ACCENT)
    c.alignment = align()
    ws.row_dimensions[row].height = size * 2


def header_row(ws, row, headers, fill_rgb=ACCENT):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = f(11, bold=True, color="FFFFFF")
        c.fill = fill(fill_rgb)
        c.alignment = align(h="center", wrap=True)
        c.border = box()
    ws.row_dimensions[row].height = 24


def note_row(ws, row, text, last_col=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = f(10, color=MUTED, italic=True)
    c.alignment = align(wrap=True)
    ws.row_dimensions[row].height = 28
    return c


def dv_list(ws, source, target, strict=False):
    dv = DataValidation(type="list", formula1=source, allow_blank=True)
    dv.showErrorMessage = strict
    ws.add_data_validation(dv)
    dv.add(target)
    return dv


def read_csv(name):
    with (DATA / name).open(encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def as_time(text):
    h, m = text.split(":")
    return time(int(h), int(m))


def data_cell(ws, row, col, value=None, editable=True, wrap=False, center=False,
              bold=False, size=11, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.border = box()
    c.alignment = align(h="center" if center else "right", wrap=wrap, v="top" if wrap else "center")
    c.font = f(size, bold=bold)
    if editable:
        c.fill = fill(INPUT_BG)
    if fmt:
        c.number_format = fmt
    return c


# ---------------------------------------------------------------------------
# כללי הלו"ז — מקור יחיד: data/schedule_template.csv
# ---------------------------------------------------------------------------
def read_schedule_template():
    return read_csv("schedule_template.csv")


def _mins(t):
    return t.hour * 60 + t.minute


def suggested_time(row, candle, havdalah):
    """הזמן המוצע לשורת תבנית, כאובייקט time (או None כשאין כניסה/צאה)."""
    base, offset = row["בסיס"].strip(), row["היסט"].strip()
    if base == "קבוע":
        return as_time(offset)
    anchor = candle if base == "כניסה" else havdalah
    if anchor is None:
        return None
    total = _mins(anchor) + int(offset)
    if int(offset):
        total = (total + 2) // 5 * 5          # עיגול ל-5 דקות, כמו ROUND בגיליון
    return time((total // 60) % 24, total % 60)


def suggestion_formula(row):
    """נוסחת ההצעה לאותה שורה, לעדכון אוטומטי כשמחליפים תאריך בשיטס."""
    base, offset = row["בסיס"].strip(), row["היסט"].strip()
    if base == "קבוע":
        return as_time(offset)
    ref = "$" + SCHED_CANDLE[0] + "$" + SCHED_CANDLE[1:] if base == "כניסה" else \
          "$" + SCHED_HAVDALAH[0] + "$" + SCHED_HAVDALAH[1:]
    minutes = int(offset)
    if minutes == 0:
        return '=IF({r}="","",{r})'.format(r=ref)
    sign = "+" if minutes > 0 else "-"
    expr = "{r}{s}TIME({h},{m},0)".format(r=ref, s=sign, h=abs(minutes) // 60, m=abs(minutes) % 60)
    return '=IF({r}="","",ROUND(({e})*288,0)/288)'.format(r=ref, e=expr)


# ---------------------------------------------------------------------------
# גיליון: הוראות
# ---------------------------------------------------------------------------
HELP_LINES = [
    ("מה עושים כל שבוע", None),
    ("1", "«לוז» — בוחרים תאריך. השעות בעמודה «הצעה» מתעדכנות לבד; עמודת «שעה» היא מה שנשלח בפועל — עורכים בה חופשי."),
    ("2", "«משימות» — לכל משימה: יום, שעה, קבוצה, כמה אנשים, ומי. עמודת «שמות» מתמלאת ע\"י tools/assign_groups.py ואפשר לתקן ידנית."),
    ("3", "«קבוצות» ו«חניכים» — מי בכל קבוצה השבת. גודל הקבוצה נגזר מהמשימות (השיא של אנשים בו-זמנית)."),
    ("4", "«מתכונים» ו«קייטרינג» — מה מכינים ומה מגיע מבחוץ לכל ארוחה. למילוי חופשי."),
    ("5", "מייצאים: python3 tools/export_pdf.py <תאריך> — לו\"ז צל (PDF) ודף שישי לכל קבוצה (PDF + PNG)."),
    ("", None),
    ("טיפים", None),
    ("•", "עוגן = האירוע בלו\"ז שלידו המשימה מופיעה בלו\"ז הצל. הרשימה הנגללת נלקחת מעמודת «אירוע» ב«לוז»."),
    ("•", "משימה בלי שעה = כולם בקבוצה עושים אותה (למשל ניקיון בסיום). משימה בלי קבוצה — מופיעה בלו\"ז הצל בלי שמות."),
    ("•", "שינית שם קבוצה? עדכן גם ב«משימות». שינית שם אירוע ב«לוז»? עדכן את העוגן במשימות שלו."),
    ("•", "ערכת בגוגל שיטס? הורד כ-xlsx אל shabbatot/<תאריך>.xlsx והרץ את הייצוא שוב — הוא קורא רק את התאים הסטטיים."),
]


def build_help(wb):
    ws = wb.create_sheet(SH_HELP)
    page(ws, tab=TAB_REF)
    widths(ws, {"A": 4, "B": 110})
    title_row(ws, 1, "הכנת שבת — מדרשת עין פרת", span="A:B", size=20)
    row = 3
    for key, text in HELP_LINES:
        if text is None:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            c = ws.cell(row=row, column=1, value=key)
            c.font = f(13, bold=True, color=ACCENT)
            c.alignment = align()
            ws.row_dimensions[row].height = 26
        else:
            k = ws.cell(row=row, column=1, value=key)
            k.font = f(11, bold=True, color=ACCENT)
            k.alignment = align(h="center", v="top")
            t = ws.cell(row=row, column=2, value=text)
            t.font = f(11)
            t.alignment = align(v="top", wrap=True)
            ws.row_dimensions[row].height = 30
        row += 1
    return ws


# ---------------------------------------------------------------------------
# גיליון: לוז — טבלה ידנית עם עמודת הצעה
# ---------------------------------------------------------------------------
def build_schedule(wb):
    ws = wb.create_sheet(SH_SCHED)
    page(ws, tab=TAB_INPUT)
    widths(ws, {"A": 12, "B": 11, "C": 30, "D": 18, "E": 34, "F": 11})
    title_row(ws, 1, "לו\"ז השבת", span="A:F", size=20)

    z = q(SH_ZMAN)
    labels = {(3, 1): "תאריך השבת", (3, 3): "פרשה", (4, 1): "כניסת שבת", (4, 3): "צאת שבת"}
    for (r, c), text in labels.items():
        lab = ws.cell(row=r, column=c, value=text)
        lab.font = f(11, bold=True)
        lab.alignment = align()
        lab.fill = fill(BAND)
        lab.border = box()
    date_cell = data_cell(ws, 3, 2, center=True, fmt="dd/mm/yyyy", size=12, bold=True)
    dv_list(ws, "{}$A$3:$A$400".format(z), SCHED_DATE, strict=True)
    parasha = data_cell(ws, 3, 4, '=IFERROR(VLOOKUP($B$3,{}$A$3:$D$400,2,FALSE),"")'.format(z),
                        editable=False)
    parasha.fill = fill(CALC_BG)
    for ref, col in ((SCHED_CANDLE, 3), (SCHED_HAVDALAH, 4)):
        c = ws[ref]
        c.value = '=IFERROR(VLOOKUP($B$3,{}$A$3:$D$400,{},FALSE),"")'.format(z, col)
        c.number_format = "hh:mm"
        c.font = f(12, bold=True, color=ACCENT)
        c.fill = fill(CALC_BG)
        c.alignment = align(h="center")
        c.border = box()
    ws.row_dimensions[3].height = ws.row_dimensions[4].height = 22
    hint = ws.cell(row=3, column=5, value="בחר תאריך מהרשימה — כניסה, צאה ופרשה נטענות מ«זמנים»")
    hint.font = f(10, color=MUTED, italic=True)
    hint.alignment = align()

    header_row(ws, SCHED_HEADER_ROW, ["יום", "שעה", "אירוע", "מקום", "הערה", "הצעה"])
    template = read_schedule_template()
    for i in range(SCHED_ROWS):
        r = SCHED_FIRST_ROW + i
        row = template[i] if i < len(template) else None
        day = row["יום"] if row else None
        data_cell(ws, r, L_DAY, day, center=True, bold=True)
        data_cell(ws, r, L_HOUR, None, center=True, bold=True, size=12, fmt="hh:mm")
        data_cell(ws, r, L_EVENT, row["אירוע"] if row else None, bold=True)
        data_cell(ws, r, L_PLACE, (row["מקום"] or None) if row else None)
        data_cell(ws, r, L_NOTE, (row["הערה"] or None) if row else None, wrap=True)
        sug = data_cell(ws, r, L_SUGGEST, suggestion_formula(row) if row else None,
                        editable=False, center=True, fmt="hh:mm")
        sug.font = f(10, color=MUTED)
        if day:
            ws.cell(row=r, column=L_DAY).fill = fill(DAY_FILLS.get(day, BAND))
        ws.row_dimensions[r].height = 22
    dv_list(ws, '"{}"'.format(DAYS), "A{}:A{}".format(SCHED_FIRST_ROW, SCHED_FIRST_ROW + SCHED_ROWS - 1))
    ws.freeze_panes = "A{}".format(SCHED_FIRST_ROW)
    note_row(ws, SCHED_FIRST_ROW + SCHED_ROWS + 1,
             "«שעה» היא הלו\"ז שנשלח בפועל — עורכים בה חופשי. «הצעה» מחושבת מכניסת/צאת השבת "
             "(הכללים ב-data/schedule_template.csv) ומתעדכנת כשמחליפים תאריך.", last_col=6)
    return ws


# ---------------------------------------------------------------------------
# גיליון: משימות — המקור היחיד לתורנויות
# ---------------------------------------------------------------------------
def build_tasks(wb):
    ws = wb.create_sheet(SH_TASKS)
    page(ws, tab=TAB_INPUT)
    widths(ws, {"A": 11, "B": 8, "C": 26, "D": 52, "E": 8, "F": 44, "G": 18, "H": 22})
    title_row(ws, 1, "משימות השבת — מי עושה מה ומתי", span="A:H", size=18)
    header_row(ws, 2, ["יום", "שעה", "קבוצה", "משימה", "אנשים", "שמות", "עוגן בלו\"ז", "הערה"])

    library = read_csv("task_library.csv")
    for i in range(TASK_ROWS):
        r = TASK_FIRST_ROW + i
        row = library[i] if i < len(library) else None
        data_cell(ws, r, T_DAY, row["יום"] if row else None, center=True)
        data_cell(ws, r, T_HOUR, as_time(row["שעה"]) if row and row["שעה"] else None,
                  center=True, bold=True, fmt="hh:mm")
        data_cell(ws, r, T_GROUP, (row["קבוצה"] or None) if row else None)
        data_cell(ws, r, T_TASK, row["משימה"] if row else None, wrap=True)
        data_cell(ws, r, T_PEOPLE, int(row["אנשים"]) if row and row["אנשים"] else None, center=True)
        names = data_cell(ws, r, T_NAMES, None, wrap=True)
        names.fill = fill(CALC_BG)
        data_cell(ws, r, T_ANCHOR, (row["עוגן"] or None) if row else None)
        data_cell(ws, r, T_NOTE, None, wrap=True)
        if row and row["יום"]:
            ws.cell(row=r, column=T_DAY).fill = fill(DAY_FILLS.get(row["יום"], BAND))
        ws.row_dimensions[r].height = 30 if row else 18

    last = TASK_FIRST_ROW + TASK_ROWS - 1
    dv_list(ws, '"{}"'.format(DAYS), "A{}:A{}".format(TASK_FIRST_ROW, last))
    dv_list(ws, "{}$A${}:$A${}".format(q(SH_GROUPS), GROUP_FIRST_ROW, GROUP_FIRST_ROW + GROUP_ROWS - 1),
            "C{}:C{}".format(TASK_FIRST_ROW, last))
    dv_list(ws, "{}$C${}:$C${}".format(q(SH_SCHED), SCHED_FIRST_ROW, SCHED_FIRST_ROW + SCHED_ROWS - 1),
            "G{}:G{}".format(TASK_FIRST_ROW, last))
    ws.freeze_panes = "A{}".format(TASK_FIRST_ROW)
    note_row(ws, last + 2,
             "«אנשים» = כמה צריך למשימה; ריק = כל הקבוצה. «שמות» מתמלא ע\"י השיבוץ ואפשר לתקן ידנית. "
             "«עוגן» = האירוע בלו\"ז שלידו המשימה תופיע בלו\"ז הצל.", last_col=8)
    return ws


# ---------------------------------------------------------------------------
# גיליון: קבוצות
# ---------------------------------------------------------------------------
def build_groups(wb):
    ws = wb.create_sheet(SH_GROUPS)
    page(ws, tab=TAB_INPUT)
    widths(ws, {"A": 28, "B": 22, "C": 16, "D": 60, "E": 8, "F": 10})
    title_row(ws, 1, "הקבוצות השבת", span="A:F", size=18)
    header_row(ws, 2, ["קבוצה", "קבוצת אם", "מוביל/ה", "חניכים", "גודל", "משימות"])

    plan = read_csv("group_plan.csv")
    t_groups = "{}$C${}:$C${}".format(q(SH_TASKS), TASK_FIRST_ROW, TASK_FIRST_ROW + TASK_ROWS - 1)
    for i in range(GROUP_ROWS):
        r = GROUP_FIRST_ROW + i
        row = plan[i] if i < len(plan) else None
        data_cell(ws, r, G_NAME, row["קבוצה"] if row else None, bold=True)
        data_cell(ws, r, G_PARENT, (row["קבוצת אם"] or None) if row else None)
        data_cell(ws, r, G_LEADER, (row["מוביל/ה"] or None) if row else None)
        members = data_cell(ws, r, G_MEMBERS, None, wrap=True)
        members.fill = fill(CALC_BG)
        size = data_cell(ws, r, G_SIZE,
                         '=IF(D{r}="","",LEN(D{r})-LEN(SUBSTITUTE(D{r},",",""))+1)'.format(r=r),
                         editable=False, center=True)
        size.fill = fill(CALC_BG)
        count = data_cell(ws, r, G_COUNT, '=IF(A{r}="","",COUNTIF({t},A{r}))'.format(r=r, t=t_groups),
                          editable=False, center=True)
        count.fill = fill(CALC_BG)
        ws.row_dimensions[r].height = 30 if row else 18
    ws.freeze_panes = "A{}".format(GROUP_FIRST_ROW)
    note_row(ws, GROUP_FIRST_ROW + GROUP_ROWS + 1,
             "המבנה הקבוע ב-data/group_plan.csv. «חניכים» נכתב ע\"י השיבוץ (מופרד בפסיקים); "
             "«גודל» ו«משימות» מחושבים.", last_col=6)
    return ws


# ---------------------------------------------------------------------------
# גיליון: חניכים
# ---------------------------------------------------------------------------
def build_students(wb):
    ws = wb.create_sheet(SH_STUDENTS)
    page(ws, tab=TAB_INPUT)
    widths(ws, {"A": 26, "B": 10, "C": 9, "D": 30, "E": 30})
    title_row(ws, 1, "חניכים — מי נמצא ומי בתורנות", span="A:E", size=18)
    header_row(ws, 2, ["שם", "תוכנית", "זמין/ה", "קבוצה השבת", "הערה"])

    students = read_csv("students.csv")
    for i in range(STUDENT_ROWS):
        r = STUDENT_FIRST_ROW + i
        row = students[i] if i < len(students) else None
        data_cell(ws, r, S_NAME, row["שם"] if row else None)
        data_cell(ws, r, S_PROGRAM, (row.get("תוכנית") or None) if row else None, center=True)
        data_cell(ws, r, S_AVAILABLE, None, center=True)
        data_cell(ws, r, S_GROUP, None)
        data_cell(ws, r, S_NOTE, None)
    last = STUDENT_FIRST_ROW + STUDENT_ROWS - 1
    dv_list(ws, '"כן,לא"', "C{}:C{}".format(STUDENT_FIRST_ROW, last))
    dv_list(ws, "{}$A${}:$A${}".format(q(SH_GROUPS), GROUP_FIRST_ROW, GROUP_FIRST_ROW + GROUP_ROWS - 1),
            "D{}:D{}".format(STUDENT_FIRST_ROW, last))
    ws.freeze_panes = "A{}".format(STUDENT_FIRST_ROW)
    return ws


# ---------------------------------------------------------------------------
# גיליונות חופשיים: מתכונים, קייטרינג
# ---------------------------------------------------------------------------
def build_recipes(wb):
    ws = wb.create_sheet(SH_RECIPES)
    page(ws, tab=TAB_INPUT)
    widths(ws, {"A": 18, "B": 26, "C": 12, "D": 50, "E": 50, "F": 24})
    title_row(ws, 1, "מתכונים — עוגות, סלטים, מטבוחה וארוחת צהריים שישי", span="A:F", size=18)
    header_row(ws, 2, ["קטגוריה", "מנה", "כמות", "מרכיבים", "הוראות", "הערה"])
    for i in range(RECIPE_ROWS):
        r = 3 + i
        data_cell(ws, r, 1, center=True)
        data_cell(ws, r, 2, bold=True)
        data_cell(ws, r, 3, center=True)
        data_cell(ws, r, 4, wrap=True)
        data_cell(ws, r, 5, wrap=True)
        data_cell(ws, r, 6, wrap=True)
    dv_list(ws, '"{}"'.format(RECIPE_KINDS), "A3:A{}".format(2 + RECIPE_ROWS))
    ws.freeze_panes = "A3"
    return ws


def build_catering(wb):
    ws = wb.create_sheet(SH_CATERING)
    page(ws, tab=TAB_INPUT)
    widths(ws, {"A": 18, "B": 34, "C": 12, "D": 44})
    title_row(ws, 1, "קייטרינג — מה מגיע לכל ארוחה", span="A:D", size=18)
    header_row(ws, 2, ["ארוחה", "מנה", "כמות", "הערה"])
    for i in range(CATERING_ROWS):
        r = 3 + i
        data_cell(ws, r, 1, center=True)
        data_cell(ws, r, 2)
        data_cell(ws, r, 3, center=True)
        data_cell(ws, r, 4, wrap=True)
    dv_list(ws, '"{}"'.format(MEALS), "A3:A{}".format(2 + CATERING_ROWS))
    ws.freeze_panes = "A3"
    return ws


# ---------------------------------------------------------------------------
# גיליון: זמנים
# ---------------------------------------------------------------------------
def build_zmanim(wb):
    ws = wb.create_sheet(SH_ZMAN)
    page(ws, tab=TAB_REF)
    widths(ws, {"A": 16, "B": 26, "C": 14, "D": 14, "E": 46})
    title_row(ws, 1, "לוח שבתות — זמני ירושלים", span="A:E", size=16)
    header_row(ws, 2, ["תאריך", "פרשה", "כניסת שבת", "צאת שבת", "ניתן לעריכה"])
    for i, row in enumerate(read_csv("zmanim.csv")):
        r = 3 + i
        d = datetime.strptime(row["תאריך"], "%d/%m/%Y").date()
        for col, val, fmt in ((1, d, "dd/mm/yyyy"), (2, row["פרשה"], None),
                              (3, as_time(row["כניסת שבת"]), "hh:mm"),
                              (4, as_time(row["צאת שבת"]), "hh:mm")):
            data_cell(ws, r, col, val, editable=(col in (3, 4)), center=(col != 2), fmt=fmt)
    ws.freeze_panes = "A3"
    hint = ws.cell(row=3, column=5,
                   value="הזמנים מחושבים לירושלים (כניסה 40 דק' לפני השקיעה, צאה 40 דק' אחריה). "
                         "אם הלוח שלך אומר אחרת — תקנו את השורה.")
    ws.merge_cells("E3:E8")
    hint.font = f(10, color=MUTED, italic=True)
    hint.alignment = align(v="top", wrap=True)
    return ws


# ---------------------------------------------------------------------------
def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_help(wb)
    build_schedule(wb)
    build_tasks(wb)
    build_groups(wb)
    build_students(wb)
    build_recipes(wb)
    build_catering(wb)
    build_zmanim(wb)
    wb.active = 1
    wb.save(OUT)
    print("נבנה: {} ({} גיליונות)".format(OUT.relative_to(ROOT), len(wb.sheetnames)))


if __name__ == "__main__":
    main()
