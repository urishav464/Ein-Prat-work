# -*- coding: utf-8 -*-
"""שיבוץ חניכים למשימות — ברמת המשימה, לא רק הקבוצה.

    python3 tools/assign_groups.py 2026-09-11
    python3 tools/assign_groups.py 2026-09-11 --dry-run

המשימות נקראות מגיליון «משימות» שבקובץ השבת (יום, שעה, קבוצה, אנשים). גודל כל
קבוצה נגזר מהן: המספר הגדול ביותר של אנשים שנדרשים בו-זמנית (שיא). אותו אדם
עושה כמה משימות בשעות שונות, ואף אחד לא מופיע בשתי משימות באותה שעה. משימה בלי
שעה = כל הקבוצה. מי שהוצמד לקבוצה ידנית (data/attendance/<תאריך>.csv, עמודת
«שיבוץ ידני») נשאר גם אם הקבוצה גדולה מהשיא.

המבנה הקבוע (קבוצות, מובילים, חברים קבועים) ב-data/group_plan.csv. כשיש יותר
זמינים ממקומות, עדיפות למי שהיה תורן פחות (data/duty_history.csv), וחניכי אלול
מתפזרים יחסית בין הקבוצות.
"""
import argparse
import csv
import random
import sys
from collections import Counter, OrderedDict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

sys.path.insert(0, str(Path(__file__).resolve().parent))
import attendance as attendance_mod
import build_workbook as bw
import roster

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
PLAN = DATA / "group_plan.csv"
DUTY = DATA / "duty_history.csv"
DAY_ORDER = {"שישי": 0, "שבת": 1, "מוצאי שבת": 2}


# ---------------------------------------------------------------------------
def read_plan():
    with PLAN.open(encoding="utf-8-sig", newline="") as fh:
        plan = []
        for row in csv.DictReader(fh):
            if not (row.get("קבוצה") or "").strip():
                continue
            plan.append({
                "name": row["קבוצה"].strip(),
                "parent": (row.get("קבוצת אם") or "").strip(),
                "leader": (row.get("מוביל/ה") or "").strip(),
                "fixed": [x.strip() for x in (row.get("חברים קבועים") or "").split(";") if x.strip()],
                "size": 0,
            })
    return plan


def duty_counts():
    if not DUTY.exists():
        return Counter()
    with DUTY.open(encoding="utf-8-sig", newline="") as fh:
        return Counter(r["שם"] for r in csv.DictReader(fh))


def record_duty(date, assignment):
    rows = []
    if DUTY.exists():
        with DUTY.open(encoding="utf-8-sig", newline="") as fh:
            rows = [r for r in csv.DictReader(fh) if r["תאריך"] != date]
    for group, members in assignment.items():
        rows += [{"תאריך": date, "שם": name, "קבוצה": group} for name in members]
    rows.sort(key=lambda r: (r["תאריך"], r["קבוצה"], r["שם"]))
    with DUTY.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["תאריך", "שם", "קבוצה"])
        writer.writeheader()
        writer.writerows(rows)


# ---------------------------------------------------------------------------
def _clean(value):
    return value.strip() if isinstance(value, str) else value


def read_tasks(path):
    """שורות גיליון «משימות» שיש בהן משימה."""
    ws = load_workbook(path)[bw.SH_TASKS]
    tasks = []
    for r in range(bw.TASK_FIRST_ROW, ws.max_row + 1):
        task = _clean(ws.cell(row=r, column=bw.T_TASK).value)
        if not task:
            continue
        hour = ws.cell(row=r, column=bw.T_HOUR).value
        if isinstance(hour, datetime):
            hour = hour.time()
        people = ws.cell(row=r, column=bw.T_PEOPLE).value
        tasks.append({
            "row": r, "day": _clean(ws.cell(row=r, column=bw.T_DAY).value) or "",
            "hour": hour or None, "group": _clean(ws.cell(row=r, column=bw.T_GROUP).value) or "",
            "task": task, "people": int(people) if people not in (None, "") else None,
        })
    return tasks


def slots_of(tasks, group):
    """קבוצות של משימות שרצות בו-זמנית: (מפתח, [משימות]) לפי סדר הזמן."""
    slots = OrderedDict()
    mine = sorted((t for t in tasks if t["group"] == group),
                  key=lambda t: (DAY_ORDER.get(t["day"], 9), t["hour"] is None, t["hour"] or datetime.min.time(), t["row"]))
    for t in mine:
        key = (t["day"], t["hour"]) if t["hour"] else ("*", t["row"])
        slots.setdefault(key, []).append(t)
    return list(slots.items())


def peak(tasks, group):
    """כמה אנשים הקבוצה צריכה: השיא של אנשים בו-זמנית."""
    return max((sum(t["people"] or 0 for t in ts) for _, ts in slots_of(tasks, group)), default=0)


def shrink_to_fit(plan, tasks, budget, fixed_sizes):
    """כשאין מספיק זמינים — מורידים אחד מהמשימות הגדולות ביותר עד שזה נכנס."""
    changed = []
    while True:
        need = sum(max(peak(tasks, g["name"]), fixed_sizes.get(g["name"], 0)) for g in plan)
        if need <= budget:
            return changed
        candidates = []
        for g in plan:
            if fixed_sizes.get(g["name"], 0) >= peak(tasks, g["name"]):
                continue                      # הקבוצה ממילא מלאה בהצמדות
            for _, ts in slots_of(tasks, g["name"]):
                if sum(t["people"] or 0 for t in ts) == peak(tasks, g["name"]):
                    candidates += [t for t in ts if (t["people"] or 0) > 1]
        if not candidates:
            return changed
        t = max(candidates, key=lambda t: (t["people"], -t["row"]))
        t.setdefault("orig", t["people"])
        t["people"] -= 1
        if t not in changed:
            changed.append(t)


# ---------------------------------------------------------------------------
def assign(plan, available, programs, past, seed, pins=None):
    """משבץ חניכים לקבוצות לפי גודל (g["size"]).

    סדר הקדימויות: הצמדות ידניות ← מובילים וחברים קבועים ← חניכי אלול (מפוזרים
    יחסית) ← השאר, כשמי שהיה תורן פחות פעמים נבחר קודם.
    """
    rng = random.Random(seed)
    groups = {g["name"]: [] for g in plan}
    taken = set()

    for name, group in (pins or {}).items():
        if group in groups and name in available and name not in taken:
            groups[group].append(name)
            taken.add(name)

    for g in plan:
        for name in ([g["leader"]] if g["leader"] else []) + g["fixed"]:
            match, _ = roster.match_name(name, available)
            if match and match not in taken:
                groups[g["name"]].append(match)
                taken.add(match)

    def pool(program):
        people = [n for n in available if n not in taken and programs.get(n) == program]
        rng.shuffle(people)
        return sorted(people, key=lambda n: past.get(n, 0))

    open_slots = lambda g: g["size"] - len(groups[g["name"]])

    elul = pool("אלול")
    placed = {g["name"]: 0 for g in plan}
    while elul:
        candidates = [g for g in plan if open_slots(g) > 0]
        if not candidates:
            break
        target = min(candidates,
                     key=lambda g: ((placed[g["name"]] + 1) / g["size"], -g["size"]))
        name = elul.pop(0)
        groups[target["name"]].append(name)
        placed[target["name"]] += 1
        taken.add(name)

    rest = pool("מדרשה") + elul
    rest.sort(key=lambda n: past.get(n, 0))
    for g in plan:
        while open_slots(g) > 0 and rest:
            name = rest.pop(0)
            groups[g["name"]].append(name)
            taken.add(name)

    for g in plan:                                     # המוביל/ה ראשון/ה ברשימה
        match, _ = roster.match_name(g["leader"], groups[g["name"]]) if g["leader"] else (None, None)
        if match:
            groups[g["name"]].remove(match)
            groups[g["name"]].insert(0, match)
    return groups


def fill_tasks(members, tasks, group):
    """מצמיד שמות לכל משימה של הקבוצה: סבב לפי סדר השעות, בלי כפילות באותה שעה."""
    names = {}
    if not members:
        return names
    pointer = 0
    for key, ts in slots_of(tasks, group):
        for t in ts:
            if t["people"] is None:
                names[t["row"]] = list(members)          # כל הקבוצה
                continue
            chosen = []
            for _ in range(min(t["people"], len(members))):
                chosen.append(members[pointer % len(members)])
                pointer += 1
            names[t["row"]] = chosen
    return names


# ---------------------------------------------------------------------------
def set_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def write_workbook(path, plan, groups, task_names, present, available):
    wb = load_workbook(path)

    ws = wb[bw.SH_TASKS]
    for r in range(bw.TASK_FIRST_ROW, ws.max_row + 1):
        set_cell(ws, r, bw.T_NAMES, ", ".join(task_names[r]) if r in task_names else None)

    ws = wb[bw.SH_GROUPS]
    for i, g in enumerate(plan):
        r = bw.GROUP_FIRST_ROW + i
        set_cell(ws, r, bw.G_NAME, g["name"])
        set_cell(ws, r, bw.G_PARENT, g["parent"] or None)
        set_cell(ws, r, bw.G_LEADER, g["leader"] or None)
        set_cell(ws, r, bw.G_MEMBERS, ", ".join(groups[g["name"]]) or None)
    for r in range(bw.GROUP_FIRST_ROW + len(plan), ws.max_row + 1):
        for col in (bw.G_NAME, bw.G_PARENT, bw.G_LEADER, bw.G_MEMBERS):
            set_cell(ws, r, col, None)

    ws = wb[bw.SH_STUDENTS]
    lookup = {name: group for group, members in groups.items() for name in members}
    for r in range(bw.STUDENT_FIRST_ROW, ws.max_row + 1):
        name = ws.cell(row=r, column=bw.S_NAME).value
        if not name:
            continue
        if name not in present:
            set_cell(ws, r, bw.S_AVAILABLE, None)
            set_cell(ws, r, bw.S_NOTE, "לא נוכח/ת")
        else:
            set_cell(ws, r, bw.S_AVAILABLE, "כן" if name in available else "לא")
            set_cell(ws, r, bw.S_NOTE, None)
        set_cell(ws, r, bw.S_GROUP, lookup.get(name))
    wb.save(path)


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="שיבוץ חניכים למשימות")
    ap.add_argument("date", help="תאריך יום שישי, למשל 2026-09-11")
    ap.add_argument("--seed", type=int, help="זרע אקראיות (לשחזור אותה חלוקה)")
    ap.add_argument("--file", help="נתיב קובץ השבת")
    ap.add_argument("--dry-run", action="store_true", help="הדפסה בלבד")
    args = ap.parse_args()

    date = datetime.strptime(args.date, "%Y-%m-%d").date()
    path = Path(args.file) if args.file else ROOT / "shabbatot" / "{}.xlsx".format(date.isoformat())
    if not path.exists():
        raise SystemExit("לא נמצא {} — הריצו קודם tools/new_shabbat.py".format(path))

    students = roster.read_students()
    programs = {s["שם"]: s.get("תוכנית", "") for s in students}
    present = attendance_mod.load(date)
    available = attendance_mod.load_available(date)
    if available is None:
        raise SystemExit("אין רשימת נוכחות ל-{} — הריצו קודם tools/attendance.py".format(args.date))
    pins = attendance_mod.load_pins(date)

    plan = read_plan()
    tasks = read_tasks(path)
    known = {g["name"] for g in plan}
    for t in tasks:
        if t["group"] and t["group"] not in known:
            print("⚠ משימה «{}» משויכת לקבוצה לא מוכרת: {}".format(t["task"], t["group"]))

    # גודל הקבוצה = שיא האנשים בו-זמנית, או מספר המוצמדים אם הוא גדול יותר
    fixed = Counter(g for n, g in pins.items() if n in available)
    for g in plan:
        for name in ([g["leader"]] if g["leader"] else []) + g["fixed"]:
            match, _ = roster.match_name(name, available)
            if match and match not in pins:
                fixed[g["name"]] += 1
    shrunk = shrink_to_fit(plan, tasks, len(available), fixed)
    for g in plan:
        g["size"] = max(peak(tasks, g["name"]), fixed.get(g["name"], 0))

    groups = assign(plan, available, programs, duty_counts(),
                    args.seed if args.seed is not None else date.toordinal(), pins=pins)
    task_names = {}
    for g in plan:
        task_names.update(fill_tasks(groups[g["name"]], tasks, g["name"]))

    placed = sum(len(v) for v in groups.values())
    print("שבת {} · {} נוכחים · {} זמינים · {} תורנים · {} קבוצות".format(
        date.strftime("%d/%m/%Y"), len(present), len(available), placed, len(plan)))
    if shrunk:
        print("⚠ לא מספיק זמינים — הוקטנו: " + "; ".join(
            "{} [{}] {}→{}".format(t["task"][:30], t["group"], t["orig"], t["people"]) for t in shrunk))

    parents = OrderedDict()
    for g in plan:
        parents.setdefault(g["parent"] or g["name"], []).append(g)
    by_row = {t["row"]: t for t in tasks}
    for parent, leaves in parents.items():
        total = sum(len(groups[g["name"]]) for g in leaves)
        print("\n■ {} ({})".format(parent, total))
        for g in leaves:
            label = g["name"].split(" · ")[-1] if " · " in g["name"] else g["name"]
            print("   {} ({}, שיא {}): {}".format(label, len(groups[g["name"]]), peak(tasks, g["name"]),
                                                 ", ".join(groups[g["name"]])))
            for _, ts in slots_of(tasks, g["name"]):
                for t in ts:
                    hour = t["hour"].strftime("%H:%M") if t["hour"] else "  —  "
                    print("      {} {} — {}".format(hour, t["task"][:44], ", ".join(task_names.get(t["row"], []))))

    loose = [t for t in tasks if not t["group"]]
    if loose:
        print("\nמשימות בלי קבוצה ({}): {}".format(len(loose), "; ".join(t["task"] for t in loose)))
    idle = [n for n in available if n not in {m for v in groups.values() for m in v}]
    print("\nלא תורנים השבת ({}): {}".format(len(idle), ", ".join(idle)))

    if args.dry_run:
        print("\n(dry-run — לא נכתב דבר)")
        return
    write_workbook(path, plan, groups, task_names, set(present), set(available))
    record_duty(date.isoformat(), groups)
    print("\nנכתב אל {}".format(path.relative_to(ROOT) if path.resolve().is_relative_to(ROOT) else path))


if __name__ == "__main__":
    main()
